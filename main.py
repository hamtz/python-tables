import os
import tkinter as tk
from tkinter import ttk
import openpyxl

def create_excel_file():
    path = "customer_data.xlsx"
    if not os.path.exists(path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Age", "Subscription", "Employment"])  # Header
        workbook.save(path)

def load_data():
    path = "customer_data.xlsx"
    if not os.path.exists(path):
        return

    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    list_values = list(sheet.values)

    if not list_values:
        return

    for col_name in list_values[0]:
        if col_name in cols:
            treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert('', tk.END, values=value_tuple)

def insert_row():
    name = name_entry.get()
    age = int(age_spinbox.get())
    subscription_status = status_combobox.get()
    employment_status = "Employed" if a.get() else "Not Employed"

    row_values = [name, age, subscription_status, employment_status]

    path = "customer_data.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    sheet.append(row_values)
    workbook.save(path)

    treeview.insert('', tk.END, values=row_values)

    name_entry.delete(0, "end")
    name_entry.insert(0, "Name")
    age_spinbox.delete(0, "end")
    age_spinbox.insert(0, "Age")

def toggle_mode():
    themes = style.theme_names()
    if mode_switch.instate(["selected"]) and "forest-light" in themes:
        style.theme_use("forest-light")
    elif "forest-dark" in themes:
        style.theme_use("forest-dark")

root = tk.Tk()

style = ttk.Style(root)
if os.path.exists("forest-light.tcl"):
    root.tk.call("source", "forest-light.tcl")
if os.path.exists("forest-dark.tcl"):
    root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

combo_list = ["Subscribed", "Not Subscribed", "Other"]
frame = ttk.Frame(root)
frame.pack()

widget_frame = ttk.Labelframe(frame, text="Insert Row")
widget_frame.grid(row=0, column=0, padx=20, pady=10)

name_entry = ttk.Entry(widget_frame)
name_entry.insert(0, "Name")
name_entry.bind("<FocusIn>", lambda e: name_entry.delete('0', 'end'))
name_entry.grid(row=0, column=0, sticky="ew", padx=5, pady=(0,5))

age_spinbox = ttk.Spinbox(widget_frame, from_=18, to=100)
age_spinbox.insert(0,"Age")
age_spinbox.grid(row=1, column=0, sticky="ew", padx=5, pady=(0,5))

status_combobox = ttk.Combobox(widget_frame, values=combo_list)
status_combobox.current(0)
status_combobox.grid(row=2, column=0, sticky="ew", padx=5, pady=(0,5))

a = tk.BooleanVar()
checkbutton = ttk.Checkbutton(widget_frame, text="Employed", variable=a)
checkbutton.grid(row=3, column=0, sticky="nsew", padx=5, pady=(0,5))

button = ttk.Button(widget_frame, text="Insert", command=insert_row)
button.grid(row=4, column=0, sticky="nsew", padx=5, pady=(0,5))

separator = ttk.Separator(widget_frame)
separator.grid(row=5, column=0, sticky="ew", padx=5, pady=(20, 10))

mode_switch = ttk.Checkbutton(widget_frame, text="Mode", style="Switch", command=toggle_mode)
mode_switch.grid(row=6, column=0, sticky="nsew", padx=5, pady=10)

treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1, pady=10)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Name", "Age", "Subscription", "Employment")
treeview = ttk.Treeview(treeFrame, show="headings", columns=cols, height=13, yscrollcommand=treeScroll.set)
treeview.column("Name", width=100)
treeview.column("Age", width=50)
treeview.column("Subscription", width=100)
treeview.column("Employment", width=100)
treeview.pack()
treeScroll.config(command=treeview.yview)

create_excel_file()
load_data()
root.mainloop()
